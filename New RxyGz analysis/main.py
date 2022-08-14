import openpyxl
import csv
import re
from collections import defaultdict
import matplotlib.pyplot as plt

# consts:
IGNORED_AA = "GIPLR"
IGNORED_AA_LIST = ["G", "I", "P", "L", "R"]
STOP_CODON = "*"


def extract_last_aa(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    first_row = True
    data = []
    for row in sheet.iter_rows(values_only=True):
        if first_row:
            first_row = False
            continue
        # Filter out empty rows
        if not all(cell is None for cell in row):
            row = row[1:row.index("*")]
            if list(row[:len(IGNORED_AA_LIST)]) == IGNORED_AA_LIST:
                row = row[len(IGNORED_AA_LIST):]

            row = list(row[-8:])
            data.append(row)
    return data


# init
res = defaultdict(lambda: defaultdict(int))
total = dict()
total_count = 0

# Calculate and extract reference data
workbook = openpyxl.load_workbook('reference_data.xlsx')
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=2, values_only=True):
    row = row[1]  # assuming first column is irrelevant
    if row:  # handling empty rows
        if row.startswith(IGNORED_AA):
            row = row[len(IGNORED_AA):]
        if STOP_CODON in row:
            row = row.split("*")[0]
        for aa in row:  # aa stands for Amino Acid
            res[aa]["ref"] += 1  # default is 0
            total_count += 1
total["ref"] = total_count
total_count = 0  # init
for aa in res.keys():
    aa_count = res[aa]
    aa_count["ref_percent"] = (aa_count["ref"] / total["ref"]) * 100
    # print(f"{aa} precent:\t{aa_count['ref_percent']}")
workbook.close()

# Calculate and extract experiment data
file_path = "experiment_data.xlsx"
sheet_name = '90mer_APPBP2 substrates'
pattern = r'R\w\wG'  # Define the pattern you want to search for
matches_filtered_c = []
matches_filtered_ab = []

result = extract_last_aa(file_path, sheet_name)
# Search for the pattern in the nested lists using list comprehension
matches = [sublist for sublist in result if re.findall(pattern, ''.join(sublist))]
for sublist in matches:
    joined_sublist = ''.join(sublist)
    matches_in_sublist = re.findall(pattern, joined_sublist)
    if matches_in_sublist:
        max_match = max(matches_in_sublist, key=lambda match: joined_sublist.index(match))
        matches_filtered_ab.append(max_match)

pattern = r'R\w\wG\w'  # to suit C location
for sublist in matches:
    joined_sublist = ''.join(sublist)
    matches_in_sublist = re.findall(pattern, joined_sublist)
    if matches_in_sublist:
        max_match = max(matches_in_sublist, key=lambda match: joined_sublist.index(match))
        matches_filtered_c.append(max_match)

for match_c in matches_filtered_c:
    aa = match_c[4]
    res[aa]["exp_c"] += 1
    total_count += 1
total["exp_c"] = total_count
total_count = 0
for match_ab in matches_filtered_ab:
    aa_a = match_ab[1]
    aa_b = match_ab[2]
    res[aa_a]["exp_a"] += 1
    res[aa_b]["exp_b"] += 1
    total_count += 1
total["exp_ab"] = total_count
total_count = 0

# Fold calculation
for aa in res.keys():
    aa_count = res[aa]
    aa_count["exp_percent_c"] = (aa_count["exp_c"] / total["exp_c"]) * 100
    aa_count["fold_c"] = aa_count["exp_percent_c"] / aa_count["ref_percent"]

    aa_count["exp_percent_b"] = (aa_count["exp_b"] / total["exp_ab"]) * 100
    aa_count["fold_b"] = aa_count["exp_percent_b"] / aa_count["ref_percent"]

    aa_count["exp_percent_a"] = (aa_count["exp_a"] / total["exp_ab"]) * 100
    aa_count["fold_a"] = aa_count["exp_percent_a"] / aa_count["ref_percent"]

# plot:
# AGGREGATED:
# Extract the keys and values for fold_c, fold_b, and fold_a
keys = res.keys()
fold_c_values = [res[key]['fold_c'] for key in keys]
fold_b_values = [res[key]['fold_b'] for key in keys]
fold_a_values = [res[key]['fold_a'] for key in keys]
# Set the x-axis values and the width of the bars
x = range(len(keys))
bar_width = 0.25
# Plot the bar graphs
plt.bar(x, fold_c_values, width=bar_width, label='rxxgX')
plt.bar(x, fold_b_values, width=bar_width, label='rxXgx', alpha=0.7)
plt.bar(x, fold_a_values, width=bar_width, label='rXxgx', alpha=0.5)
# Set the x-axis ticks and labels
plt.xticks(x, keys)
# Set the labels and title
plt.xlabel('Amino Acids')
plt.ylabel('Fold Values')
plt.title('Fold as function of AA - Aggregated View')
# Add a legend
plt.legend()
# Show the plot
plt.show()
plt.savefig("AggregatedView.png")

# per location:
# Extract the keys and values for fold_c, fold_b, and fold_a
keys = list(res.keys())
fold_c_values = [res[key]['fold_c'] for key in keys]
fold_b_values = [res[key]['fold_b'] for key in keys]
fold_a_values = [res[key]['fold_a'] for key in keys]
# Set the x-axis values and the width of the bars
x = range(len(keys))
bar_width = 0.35
# Plot the bar graph for Fold C
plt.bar(x, fold_c_values, width=bar_width)
plt.xlabel('Amino Acids')
plt.ylabel('Fold Values')
plt.title('Fold as function of AA - Location 0 (rxxgX)')
plt.xticks(x, keys)
plt.show()
plt.savefig("rxxgX.png")

# Plot the bar graph for Fold B
plt.bar(x, fold_b_values, width=bar_width)
plt.xlabel('Amino Acids')
plt.ylabel('Fold Values')
plt.title('Fold as function of AA - Location -2 (rxXgx)')
plt.xticks(x, keys)
plt.show()
plt.savefig("rxXgx.png")

# Plot the bar graph for Fold A
plt.bar(x, fold_a_values, width=bar_width)
plt.xlabel('Amino Acids')
plt.ylabel('Fold Values')
plt.title('Fold as function of AA - Location -3 (rXxgx)')
plt.xticks(x, keys)
plt.show()
plt.savefig("rXxgx.png")

# Export CSV file
csv_file = 'outputs_data.csv'

def write_dict_to_csv(dictionary, filename):
    with open(filename, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        write_dict_recursive(dictionary, writer)

def write_dict_recursive(dictionary, writer, parent_key=''):
    for key, value in dictionary.items():
        if isinstance(value, dict):
            new_parent_key = f'{parent_key}.{key}' if parent_key else key
            write_dict_recursive(value, writer, new_parent_key)
        else:
            fieldname = f'{parent_key}.{key}' if parent_key else key
            writer.writerow([fieldname, value])

# Example usage
write_dict_to_csv(res, csv_file)
